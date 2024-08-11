import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Grades"

#Column Names 
headers = ["Student Names", "Math", "Science", "English", "Gym"]
data = [
    ["Joe", 85, 92, 78, 95],
    ["Bill", 92, 88, 85, 87],
    ["Tim", 78, 80, 92, 90],
    ["Sally", 95, 85, 88, 82],
    ["Jane", 88, 91, 87, 93]
]

# Write headers
for col, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Write data
for row, student_data in enumerate(data, start=2):
    for col, value in enumerate(student_data, start=1):
        sheet.cell(row=row, column=col, value=value)

# Calculate averages
last_row = len(data) + 2
for col in range(2, 6):  # Columns B to E
    sheet.cell(row=last_row, column=col, value=f"=AVERAGE({chr(64+col)}2:{chr(64+col)}{last_row-1})")

# Save the workbook
wb.save("student_grades.xlsx")